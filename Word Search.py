# To work this code properly, install and import necessary modules.

import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys             # This will alow us to use 'Enter'/'ESC' (any Keyboard keys) in automation.
from selenium.webdriver.common.by import By                 # Without this, [By] in [browser.find_element(By.CSS_SELECTOR] will not work.
# import time                                                 # This is for using delays in this code.
import openpyxl                                             # This is for working on excel file.
from selenium.common.exceptions import NoSuchWindowException, NoSuchElementException # This is for handling exceptions when browser is closed.
# import win32com.client as win32                             # [pip install pywin32] command is needed for this line of code.
import comtypes.client
import ChangePart                                           # This will import File Names and Locations which must be changed from system to system.
import Additional


####################        These set of code is for quiting Excel Aplication if already opened.
# xlApp = win32.gencache.EnsureDispatch('Excel.Application')  # Quits Excel app if already open else skips the process (if not opened).
xlApp = comtypes.client.CreateObject('Excel.Application')
# try:                                                        # This [try & except] clause saves + closes excel app if not saved else closes excel if no need of 'save'.
#     xlApp.ActiveWorkbook.Save()
#     xlApp.Application.Quit() 
# except:
#     xlApp.Application.Quit() 

####################        These sets of code is for opening chrome browser.
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)
os.environ['PATH'] += ChangePart.ChromeLocation
driver.maximize_window()

####################        Below code will open excel file and read/write data on it.
xlFile = ChangePart.xlFileName                              # 'r' converts string to a 'raw string'.
workbook = openpyxl.load_workbook(filename= xlFile)
# worksheet = workbook[ChangePart.activeWorkSheet]
worksheet = workbook.active

###################################################
def bdword():
    driver.get("https://www.bdword.com/english-to-bengali-meaning-" + word)
    srch_bdword = driver.find_element(By.CSS_SELECTOR, "div.align_text2")         # We must have to use 'div' in ["div.align_text2"] which we got on hovering over the class name.
    cell.offset(row = 0, column = ChangePart.offset_output).value = srch_bdword.text                     # [.offset] method will move to next Column, [.value] will paste the word to next column, [.text] will convert searched word to text.

def eng2ban():
    driver.get("https://www.english-bangla.com/dictionary/" + word)
    srch_bdword = driver.find_element(By.CSS_SELECTOR, "span.format1")
    cell.offset(row = 0, column = ChangePart.offset_output).value = srch_bdword.text
    
def OED():
    driver.get("https://www.oed.com/search/dictionary/?scope=Entries&q=" + word)
    srch_bdword = driver.find_element(By.CSS_SELECTOR, "div.snippet")
    cell.offset(row = 0, column = ChangePart.offset_output).value = srch_bdword.text
    
def Merrium():
    driver.get("https://www.merriam-webster.com/dictionary/" + word)
    srch_bdword = driver.find_element(By.CSS_SELECTOR, "div.vg")
    cell.offset(row = 0, column = ChangePart.offset_output).value = srch_bdword.text
    
def collings():
    driver.get("https://www.collinsdictionary.com/dictionary/english/" + word)
    srch_bdword = driver.find_element(By.CSS_SELECTOR, "div.hom")
    cell.offset(row = 0, column = ChangePart.offset_output).value = srch_bdword.text
    
###################################################
try:
    for row in worksheet.iter_rows(min_row=ChangePart.row_from, max_row=ChangePart.row_to, min_col=ChangePart.col_from, max_col=ChangePart.col_to, values_only=False):
        # Here, iter_rows() method returns a generator that iterates over all the rows in the worksheet. 
        # You can specify the range of rows and columns to iterate over using the min_row, max_row, min_col, and max_col parameters.
        for cell in row:
            word =  cell.value
            try:
                # bdword()
                eng2ban()
                # OED()
                # Merrium()
                # collings()
            except NoSuchElementException:
                print(f'Word "{word}" is not found in this dictionary.')
                # Here the loop will not break.
                
            except NoSuchWindowException:
                print("Browser window was closed by clicking.")
                break  # Exit the loop if browser is closed
            
            except Exception as e:
                print(f"An error occurred: {e}")
                break  # Exit the loop on any other exception
            
finally:
    driver.close()

####################     Saves and closes the workbook which is opened by the driver.
workbook.save(xlFile)
workbook.close()

####################     Opens the excel workbook in user-view mode.
open_wb = xlApp.Workbooks.Open(xlFile)
opnxl = open_wb.Worksheets(ChangePart.activeWorkSheet) 
xlApp.Visible = True
