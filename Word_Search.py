
import openpyxl                                                                                    # This is for working on excel file.
from selenium.common.exceptions import NoSuchWindowException, NoSuchElementException               # This is for handling exceptions and errors (More Specifically when browser closed).

import ChangePart
import Additional

####################        Below code is for selecting the 'Excel application'
xlApp = Additional.xlApp

####################        These sets of code is for opening chrome browser.
driver = ChangePart.Driver_Select
driver.maximize_window()

####################        Below code will open excel file and read/write data on it.
xlFile = ChangePart.xlFileName                              
workbook = openpyxl.load_workbook(filename= xlFile)
# worksheet = workbook[ChangePart.activeWorkSheet]
worksheet = workbook.active

####################
try:
    for row in worksheet.iter_rows(min_row=ChangePart.row_from, max_row=ChangePart.row_to, min_col=ChangePart.col_from, max_col=ChangePart.col_to, values_only=False):
        # Here, iter_rows() method returns a generator that iterates over all the rows(from left to right) in the worksheet. 
        # You can specify the range of rows and columns to iterate over using the [min_row, max_row, min_col, max_col, values_only] parameters.
        for cell in row:
            try:
                Dict_word = ChangePart.dictionary_name(cell.value)
                cell.offset(row = 0, column = ChangePart.offset_output).value = Dict_word.text      # [.offset] method will move to next Column, [.value] will paste the word to next column, [.text] will convert searched word to text.
                
            except NoSuchElementException:
                print(f'Word "{cell.value}" is not found in this dictionary.')
                # Here the loop will not break.
                
            except NoSuchWindowException:
                print("Browser window was closed by clicking.")
                break  # Exit the loop if browser is closed
            
            except Exception as e:
                print(f"An error occurred: {e}")
                break  # Exit the loop on any other exception
            
finally:
    try:
        driver.close()
    except:
        print("Browser window already closed by clicking.")

####################     Saves and closes the workbook which is opened by the driver.
try:
    workbook.save(xlFile)
    workbook.close()
except Exception as e:
    print(f"An error occurred: {e}")

####################     Opens the excel workbook in user-view mode.
open_wb = xlApp.Workbooks.Open(xlFile)
opnxl = open_wb.Worksheets(ChangePart.activeWorkSheet)
xlApp.Visible = True