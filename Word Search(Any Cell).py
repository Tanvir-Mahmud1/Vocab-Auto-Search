# These codes needs moderation as per your need. So go through the code and and modify it accordingly.

# To work this code properly, install and import bellow modules and [Very Very Important]=> Check "Chrome Driver" and "word" file path.
# pip install openpyxl, pywin32


import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
# This will alow us to use 'Enter' or 'ESC' keys in automation process.
from selenium.webdriver.common.by import By
import time
import openpyxl
#This is for working on excel file.

# This set of code is for quiting Excel Aplication if already opened.
# But to run [win32com.client] we first need to install [pip install pywin32] in command prompt.
import win32com.client as win32
# Quit the Excel application if already opened or else will skip the process if not opened. Exactly these two line of code is needed to do so.
xlApp = win32.gencache.EnsureDispatch('Excel.Application')
xlApp.Application.Quit()

# This set of code is for opening chrome browser.
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)
# Check PATH Bellow......................................................................................................................
os.environ['PATH'] += r"C:/Chrome Driver"

# Check xlfile Name Bellow......................................................................................................................
xlfile = r"D:\Projects\Vocab-Auto-Search\1500_Words.xlsx"
# [r''] converts string to a raw string.
workbook = openpyxl.load_workbook(filename= xlfile)
worksheet = workbook.active
# or
# worksheet = workbook['Sheet1']

for row in worksheet.iter_rows():
# Here, iter_rows() method returns a generator that iterates over all the rows in the worksheet. 
# You can specify the range of rows and columns to iterate over using the min_row, max_row, min_col, and max_col parameters.
    for cell in row:

        word = cell.value
        driver.maximize_window()
        driver.get("https://www.bdword.com/english-to-bengali-meaning-" + word)
        # print(cell.value)

        srcWord = driver.find_element(By.CSS_SELECTOR, "div.align_text2")
        # We must have to use 'div' in ["div.align_text2"] which we got on hovering over the class name (which we want) while inspecting the browser content.
        resultWord = srcWord.text
        print(resultWord)


driver.close()

workbook.save(xlfile)
workbook.close()

# These lines of code will open WorkBook (+WorkSheet) where searched data was saved.
open_wb = xlApp.Workbooks.Open(xlfile)
# Check Sheet Name Bellow......................................................................................................................
opnxl = open_wb.Worksheets('1500 Words')
xlApp.Visible = True